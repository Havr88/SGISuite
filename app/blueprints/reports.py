from flask import Blueprint, render_template, request, make_response
from flask_login import login_required, current_user
from app.models import db, InventoryMovement, InstitutionConfig, Article
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

reports = Blueprint('reports', __name__)

@reports.route('/reports/movements')
@login_required
def movements():
    from datetime import datetime
    from sqlalchemy import extract
    from app.models import Article, Department
    
    # Obtener filtros de la URL
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    article_id = request.args.get('article_id', type=int)
    department_id = request.args.get('department_id', type=int)
    
    query = InventoryMovement.query
    
    if year:
        query = query.filter(extract('year', InventoryMovement.date) == year)
    if month:
        query = query.filter(extract('month', InventoryMovement.date) == month)
    if article_id:
        query = query.filter(InventoryMovement.article_id == article_id)
    if department_id:
        query = query.filter(InventoryMovement.department_id == department_id)
    
    page = request.args.get('page', 1, type=int)
    pagination = query.order_by(InventoryMovement.date.desc()).paginate(page=page, per_page=15, error_out=False)
    articles = Article.query.all()
    departments = Department.query.all()
    
    return render_template('report_movements.html', 
                           pagination=pagination,
                           movements=pagination.items,
                           articles=articles,
                           departments=departments,
                           current_month=month,
                           current_year=year,
                           current_article_id=article_id,
                           current_department_id=department_id)

def _get_movement_data_and_config(month=None, year=None, article_id=None, department_id=None):
    from sqlalchemy import extract
    query = InventoryMovement.query
    if year:
        query = query.filter(extract('year', InventoryMovement.date) == year)
    if month:
        query = query.filter(extract('month', InventoryMovement.date) == month)
    if article_id:
        query = query.filter(InventoryMovement.article_id == article_id)
    if department_id:
        query = query.filter(InventoryMovement.department_id == department_id)
        
    movements_data = query.order_by(InventoryMovement.date.desc()).all()
    config = InstitutionConfig.query.first()
    return movements_data, config

@reports.route('/reports/movements/pdf')
@login_required
def download_movements_pdf():
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    article_id = request.args.get('article_id', type=int)
    department_id = request.args.get('department_id', type=int)
    
    movements_data, config = _get_movement_data_and_config(month, year, article_id, department_id)
    inst_name = config.name if config else "Institución"
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1 # Center
    
    # Header
    elements.append(Paragraph(f"<b>{inst_name}</b>", title_style))
    elements.append(Paragraph("Reporte de Movimientos de Inventario (SGI-Docs+)", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table Data
    data = [["Fecha", "Tipo", "Artículo", "Cantidad", "Usuario/Destino"]]
    for m in movements_data:
        date_str = m.date.strftime('%d/%m/%Y %H:%M')
        m_type = "INGRESO" if m.movement_type == 'IN' else "ENTREGA"
        art_name = m.article.name if m.article else 'N/A'
        
        # Uso de unidades técnicas
        qty = f"{m.quantity} {m.article.base_unit.abbreviation if m.article and m.article.base_unit else ''}"
        
        # Lógica de destino detallada para egresos
        if m.movement_type == 'OUT':
            dest_parts = []
            if m.department: dest_parts.append(m.department.name)
            if m.receiver_name: dest_parts.append(f"Recibe: {m.receiver_name}")
            if m.receiver_cedula: dest_parts.append(f"ID: {m.receiver_cedula}")
            dest = "\n".join(dest_parts) if dest_parts else "N/A"
        else:
            dest = f"Carga: {m.user.username}" if m.user else "Sistema"
            
        data.append([date_str, m_type, art_name, qty, dest])
        
    # Table Style
    t = Table(data, colWidths=[100, 70, 150, 70, 150])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(t)
    doc.build(elements)
    
    pdf_out = buffer.getvalue()
    buffer.close()
    
    response = make_response(pdf_out)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_movimientos.pdf'
    return response

@reports.route('/reports/correlation')
@login_required
def correlation():
    from datetime import datetime
    import calendar
    from app.models import Article
    
    month = request.args.get('month', type=int, default=datetime.now().month)
    year = request.args.get('year', type=int, default=datetime.now().year)
    
    # Fecha de inicio y fin del mes seleccionado
    last_day = calendar.monthrange(year, month)[1]
    start_date = datetime(year, month, 1)
    end_date = datetime(year, month, last_day, 23, 59, 59)
    
    articles = Article.query.all()
    correlation_data = []
    
    for art in articles:
        # Stock al final del mes = Stock Actual - Net(movimientos después de end_date)
        net_after = db.session.query(
            db.func.sum(db.case((InventoryMovement.movement_type == 'IN', InventoryMovement.quantity), else_=-InventoryMovement.quantity))
        ).filter(InventoryMovement.article_id == art.id, InventoryMovement.date > end_date).scalar() or 0
        
        closing_stock = art.current_stock - net_after
        
        # Movimientos durante el mes
        ins_month = db.session.query(db.func.sum(InventoryMovement.quantity)).filter(
            InventoryMovement.article_id == art.id,
            InventoryMovement.movement_type == 'IN',
            InventoryMovement.date >= start_date,
            InventoryMovement.date <= end_date
        ).scalar() or 0
        
        outs_month = db.session.query(db.func.sum(InventoryMovement.quantity)).filter(
            InventoryMovement.article_id == art.id,
            InventoryMovement.movement_type == 'OUT',
            InventoryMovement.date >= start_date,
            InventoryMovement.date <= end_date
        ).scalar() or 0
        
        opening_stock = closing_stock - ins_month + outs_month
        
        correlation_data.append({
            'article': art,
            'opening': opening_stock,
            'ins': ins_month,
            'outs': outs_month,
            'closing': closing_stock
        })
        
    return render_template('inventory_correlation.html',
                           data=correlation_data,
                           current_month=month,
                           current_year=year)

@reports.route('/reports/articles/pdf')
@login_required
def download_articles_pdf():
    from app.models import Article
    articles = Article.query.all()
    config = InstitutionConfig.query.first()
    inst_name = config.name if config else "Institución"
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.append(Paragraph(f"<b>{inst_name}</b>", styles['Heading1']))
    elements.append(Paragraph("Inventario General de Artículos", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    data = [["Artículo", "Categoría", "Ubicación", "Stock Actual", "Mínimo"]]
    for a in articles:
        data.append([
            a.name, 
            a.category.name, 
            a.location or '-', 
            f"{a.current_stock} {a.base_unit.abbreviation}",
            f"{a.min_stock} {a.base_unit.abbreviation}"
        ])
    
    t = Table(data, colWidths=[180, 100, 100, 80, 60])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(t)
    doc.build(elements)
    
    pdf_out = buffer.getvalue()
    buffer.close()
    
    response = make_response(pdf_out)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=inventario_general.pdf'
    return response

@reports.route('/reports/correlation/pdf')
@login_required
def download_correlation_pdf():
    from app.models import Article
    import calendar
    month = request.args.get('month', type=int, default=datetime.now().month)
    year = request.args.get('year', type=int, default=datetime.now().year)
    
    # Reutilizar lógica de correlación (simplificada para PDF)
    last_day = calendar.monthrange(year, month)[1]
    start_date = datetime(year, month, 1)
    end_date = datetime(year, month, last_day, 23, 59, 59)
    
    articles = Article.query.all()
    config = InstitutionConfig.query.first()
    inst_name = config.name if config else "Institución"
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    month_name = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"][month-1]
    elements.append(Paragraph(f"<b>{inst_name}</b>", styles['Heading1']))
    elements.append(Paragraph(f"Auditoría de Correlación Mensual: {month_name} {year}", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    data = [["Artículo", "Saldo Inicial", "(+) Ingresos", "(-) Egresos", "Saldo Final"]]
    for art in articles:
        # Lógica de cálculo (idéntica a la vista correlation)
        net_after = db.session.query(
            db.func.sum(db.case((InventoryMovement.movement_type == 'IN', InventoryMovement.quantity), else_=-InventoryMovement.quantity))
        ).filter(InventoryMovement.article_id == art.id, InventoryMovement.date > end_date).scalar() or 0
        closing_stock = art.current_stock - net_after
        ins_month = db.session.query(db.func.sum(InventoryMovement.quantity)).filter(
            InventoryMovement.article_id == art.id, InventoryMovement.movement_type == 'IN',
            InventoryMovement.date >= start_date, InventoryMovement.date <= end_date
        ).scalar() or 0
        outs_month = db.session.query(db.func.sum(InventoryMovement.quantity)).filter(
            InventoryMovement.article_id == art.id, InventoryMovement.movement_type == 'OUT',
            InventoryMovement.date >= start_date, InventoryMovement.date <= end_date
        ).scalar() or 0
        opening_stock = closing_stock - ins_month + outs_month
        
        data.append([
            art.name,
            f"{opening_stock:.3f}",
            f"{ins_month:.3f}",
            f"{outs_month:.3f}",
            f"{closing_stock:.3f}"
        ])
    
    t = Table(data, colWidths=[180, 80, 80, 80, 80])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')), # Purple for Audit
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
    ]))
    
    elements.append(t)
    doc.build(elements)
    
    pdf_out = buffer.getvalue()
    buffer.close()
    
    response = make_response(pdf_out)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=auditoria_{month}_{year}.pdf'
    return response

@reports.route('/reports/movements/xlsx')
@login_required
def download_movements_xlsx():
    month = request.args.get('month', type=int)
    year = request.args.get('year', type=int)
    article_id = request.args.get('article_id', type=int)
    department_id = request.args.get('department_id', type=int)
    
    movements_data, config = _get_movement_data_and_config(month, year, article_id, department_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    
    # Header
    headers = ["Fecha", "Tipo", "Artículo", "Cantidad", "Unidad", "Usuario/Carga", "Destino/Departamento", "Cédula Receptor", "Observaciones"]
    ws.append(headers)
    
    # Style Header
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for m in movements_data:
        m_type = "INGRESO" if m.movement_type == 'IN' else "ENTREGA"
        dest = m.department.name if m.department else "N/A"
        ws.append([
            m.date.strftime('%d/%m/%Y %H:%M'),
            m_type,
            m.article.name if m.article else 'N/A',
            m.quantity,
            m.article.base_unit.abbreviation if m.article and m.article.base_unit else '',
            m.user.username if m.user else "Sistema",
            dest,
            m.receiver_cedula or "N/A",
            m.observations or ""
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    xlsx_out = buffer.getvalue()
    buffer.close()
    
    response = make_response(xlsx_out)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_movimientos.xlsx'
    return response

@reports.route('/reports/articles/xlsx')
@login_required
def download_articles_xlsx():
    articles = Article.query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario General"
    
    headers = ["Artículo", "Categoría", "Ubicación", "Stock Actual", "Unidad", "Costo Unitario", "Valor Total", "Stock Mínimo"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for a in articles:
        ws.append([
            a.name,
            a.category.name,
            a.location or "",
            a.current_stock,
            a.base_unit.abbreviation,
            a.unit_cost or 0,
            (a.current_stock or 0) * (a.unit_cost or 0),
            a.min_stock
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    xlsx_out = buffer.getvalue()
    buffer.close()
    
    response = make_response(xlsx_out)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=inventario_general.xlsx'
    return response
